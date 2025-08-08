from typing import List, Optional
from datetime import datetime, timedelta
from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, or_, desc
from sqlalchemy.orm import selectinload, joinedload
from sqlalchemy.exc import IntegrityError

from app.api.deps import get_async_session, get_current_admin
from app.db.models.product import Category, Product, ProductPackage, District, PromoCode
from app.db.models.user import User
from app.db.models.order import Order, OrderItem
from app.db.models.admin import AdminUser
from app.schemas.product import (
    Category as CategorySchema, 
    Product as ProductSchema,
    ProductPackage as ProductPackageSchema,
    ProductPackageCreate,
    ProductPackageUpdate
)
from app.schemas.admin import *
from app.services.s3 import s3_service

router = APIRouter()

# File Upload
@router.post("/upload/image", response_model=ImageUploadResponse)
async def upload_image(
    image: UploadFile = File(...),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Upload image to S3"""
    try:
        image_url = await s3_service.upload_image(image, folder="products")
        return ImageUploadResponse(url=image_url)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Categories CRUD
@router.get("/categories", response_model=List[CategorySchema])
async def get_admin_categories(
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get all categories for admin"""
    query = select(Category).order_by(Category.order)
    result = await session.execute(query)
    categories = result.scalars().all()
    return categories

@router.post("/categories", response_model=CategorySchema)
async def create_category(
    category: CategoryCreate,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Create new category"""
    try:
        db_category = Category(**category.model_dump())
        session.add(db_category)
        await session.commit()
        await session.refresh(db_category)
        return db_category
    except IntegrityError:
        await session.rollback()
        raise HTTPException(status_code=400, detail="Category with this ID already exists")

@router.put("/categories/{category_id}", response_model=CategorySchema)
async def update_category(
    category_id: str,
    category: CategoryUpdate,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Update category"""
    query = select(Category).where(Category.id == category_id)
    result = await session.execute(query)
    db_category = result.scalar_one_or_none()
    
    if not db_category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    for field, value in category.model_dump(exclude_unset=True).items():
        setattr(db_category, field, value)
    
    await session.commit()
    await session.refresh(db_category)
    return db_category

@router.delete("/categories/{category_id}")
async def delete_category(
    category_id: str,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Delete category"""
    query = select(Category).where(Category.id == category_id)
    result = await session.execute(query)
    db_category = result.scalar_one_or_none()
    
    if not db_category:
        raise HTTPException(status_code=404, detail="Category not found")
    
    await session.delete(db_category)
    await session.commit()
    return {"message": "Category deleted successfully"}

# Products CRUD
@router.get("/products", response_model=PaginatedResponse[ProductResponse])
async def get_admin_products(
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get paginated products for admin"""
    offset = (page - 1) * size
    
    # Count total
    count_query = select(func.count(Product.id))
    count_result = await session.execute(count_query)
    total = count_result.scalar()
    
    # Get products
    query = (
        select(Product)
        .options(
            selectinload(Product.category),
            selectinload(Product.product_packages)
        )
        .order_by(desc(Product.is_featured), Product.name)
        .offset(offset)
        .limit(size)
    )
    result = await session.execute(query)
    products = result.scalars().all()
    
    return PaginatedResponse(
        items=products,
        total=total,
        page=page,
        size=size
    )

@router.get("/products/stats")
async def get_product_stats(
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get product statistics"""
    total_products_query = select(func.count(Product.id))
    total_categories_query = select(func.count(Category.id))
    featured_products_query = select(func.count(Product.id)).where(Product.is_featured == True)
    active_products_query = select(func.count(Product.id)).where(Product.is_active == True)
    
    total_products = (await session.execute(total_products_query)).scalar()
    total_categories = (await session.execute(total_categories_query)).scalar()
    featured_products = (await session.execute(featured_products_query)).scalar()
    active_products = (await session.execute(active_products_query)).scalar()
    
    return {
        "total_products": total_products,
        "total_categories": total_categories,
        "featured_products": featured_products,
        "active_products": active_products
    }

@router.get("/products/{product_id}", response_model=ProductResponse)
async def get_admin_product(
    product_id: str,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get product by ID"""
    query = select(Product).options(
        selectinload(Product.category),
        selectinload(Product.product_packages)
    ).where(Product.id == product_id)
    result = await session.execute(query)
    product = result.scalar_one_or_none()
    
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    return product

@router.post("/products", response_model=ProductResponse)
async def create_product(
    product: ProductCreate,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Create new product"""
    try:
        db_product = Product(**product.model_dump())
        session.add(db_product)
        await session.commit()
        await session.refresh(db_product)
        
        # Load category
        await session.refresh(db_product, ['category'])
        return db_product
    except IntegrityError:
        await session.rollback()
        raise HTTPException(status_code=400, detail="Product with this ID already exists or invalid category")

@router.put("/products/{product_id}", response_model=ProductResponse)
async def update_product(
    product_id: str,
    product: ProductUpdate,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Update product"""
    query = select(Product).where(Product.id == product_id)
    result = await session.execute(query)
    db_product = result.scalar_one_or_none()
    
    if not db_product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Delete old image if new one is provided
    if product.image_url and db_product.image_url and product.image_url != db_product.image_url:
        await s3_service.delete_image(db_product.image_url)
    
    for field, value in product.model_dump(exclude_unset=True).items():
        setattr(db_product, field, value)
    
    await session.commit()
    await session.refresh(db_product, ['category'])
    return db_product

@router.delete("/products/{product_id}")
async def delete_product(
    product_id: str,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Delete product"""
    query = select(Product).where(Product.id == product_id)
    result = await session.execute(query)
    db_product = result.scalar_one_or_none()
    
    if not db_product:
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Delete image from S3
    if db_product.image_url:
        await s3_service.delete_image(db_product.image_url)
    
    await session.delete(db_product)
    await session.commit()
    return {"message": "Product deleted successfully"}


# Users Management
@router.get("/users", response_model=PaginatedResponse[UserResponse])
async def get_admin_users(
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get paginated users for admin"""
    offset = (page - 1) * size
    
    # Count total
    count_query = select(func.count(User.id))
    count_result = await session.execute(count_query)
    total = count_result.scalar()
    
    # Get users
    query = (
        select(User)
        .order_by(desc(User.created_at))
        .offset(offset)
        .limit(size)
    )
    result = await session.execute(query)
    users = result.scalars().all()
    
    return PaginatedResponse(
        items=users,
        total=total,
        page=page,
        size=size
    )

@router.get("/users/stats", response_model=UserStats)
async def get_user_stats(
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get user statistics"""
    total_query = select(func.count(User.id))
    active_query = select(func.count(User.id)).where(User.is_blocked == False)
    gold_query = select(func.count(User.id)).where(User.is_gold_client == True)
    blocked_query = select(func.count(User.id)).where(User.is_blocked == True)
    
    total = (await session.execute(total_query)).scalar()
    active = (await session.execute(active_query)).scalar()
    gold_clients = (await session.execute(gold_query)).scalar()
    blocked = (await session.execute(blocked_query)).scalar()
    
    return UserStats(
        total=total,
        active=active,
        gold_clients=gold_clients,
        blocked=blocked
    )

@router.get("/users/{user_id}", response_model=UserResponse)
async def get_user_by_id(
    user_id: int,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get user by ID"""
    query = select(User).where(User.id == user_id)
    result = await session.execute(query)
    user = result.scalar_one_or_none()
    
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    return user

@router.put("/users/{user_id}", response_model=UserResponse)
async def update_user(
    user_id: int,
    user: UserUpdate,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Update user"""
    query = select(User).where(User.id == user_id)
    result = await session.execute(query)
    db_user = result.scalar_one_or_none()
    
    if not db_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    for field, value in user.model_dump(exclude_unset=True).items():
        setattr(db_user, field, value)
    
    await session.commit()
    await session.refresh(db_user)
    return db_user

# Orders Management
@router.get("/orders", response_model=PaginatedResponse[OrderResponse])
async def get_admin_orders(
    page: int = Query(1, ge=1),
    size: int = Query(20, ge=1, le=100),
    status: Optional[str] = Query(None),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    order_id: Optional[int] = Query(None),
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get paginated orders for admin"""
    offset = (page - 1) * size
    
    # Build filters
    filters = []
    if status:
        from app.db.models.order import OrderStatus
        try:
            status_enum = OrderStatus(status)
            filters.append(Order.status == status_enum)
        except ValueError:
            # Invalid status, no results
            filters.append(Order.id == -1)
    if order_id:
        filters.append(Order.order_id == order_id)
    if start_date:
        try:
            start_dt = datetime.fromisoformat(start_date)
            filters.append(Order.created_at >= start_dt)
        except ValueError:
            pass
    if end_date:
        try:
            end_dt = datetime.fromisoformat(end_date)
            filters.append(Order.created_at <= end_dt)
        except ValueError:
            pass
    
    # Count total
    count_query = select(func.count(Order.id))
    if filters:
        count_query = count_query.where(and_(*filters))
    count_result = await session.execute(count_query)
    total = count_result.scalar()
    
    # Get orders
    query = (
        select(Order)
        .options(
            selectinload(Order.items),
            selectinload(Order.user),
            selectinload(Order.district)
        )
        .order_by(desc(Order.created_at))
        .offset(offset)
        .limit(size)
    )
    
    if filters:
        query = query.where(and_(*filters))
    
    result = await session.execute(query)
    orders = result.scalars().all()
    
    return PaginatedResponse(
        items=orders,
        total=total,
        page=page,
        size=size
    )

@router.get("/orders/stats", response_model=OrderStats)
async def get_order_stats(
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get order statistics"""
    # Build filters
    filters = []
    if start_date:
        try:
            start_dt = datetime.fromisoformat(start_date)
            filters.append(Order.created_at >= start_dt)
        except ValueError:
            pass
    if end_date:
        try:
            end_dt = datetime.fromisoformat(end_date)
            filters.append(Order.created_at <= end_dt)
        except ValueError:
            pass
    
    # Base query with filters
    base_query = select(Order)
    if filters:
        base_query = base_query.where(and_(*filters))
    
    # Total orders
    total_orders_query = select(func.count(Order.id))
    if filters:
        total_orders_query = total_orders_query.where(and_(*filters))
    total_orders = (await session.execute(total_orders_query)).scalar()
    
    # Total revenue
    revenue_query = select(func.sum(Order.total_amount))
    if filters:
        revenue_query = revenue_query.where(and_(*filters))
    total_revenue = (await session.execute(revenue_query)).scalar() or 0
    
    # Average order value
    avg_order_value = total_revenue / total_orders if total_orders > 0 else 0
    
    # Orders by status
    status_query = select(Order.status, func.count(Order.id)).group_by(Order.status)
    if filters:
        status_query = status_query.where(and_(*filters))
    status_result = await session.execute(status_query)
    orders_by_status = {status.value if hasattr(status, 'value') else str(status): count 
                       for status, count in status_result.fetchall()}
    
    return OrderStats(
        total_orders=total_orders,
        total_revenue=total_revenue,
        avg_order_value=avg_order_value,
        orders_by_status=orders_by_status
    )

@router.get("/orders/export")
async def export_orders_report(
    start_date: str = Query(...),
    end_date: str = Query(...),
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Export orders report as Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        from fastapi.responses import StreamingResponse
        
        start_dt = datetime.fromisoformat(start_date)
        end_dt = datetime.fromisoformat(end_date)
        
        # Get orders
        query = (
            select(Order)
            .options(
                selectinload(Order.items),
                selectinload(Order.user),
                selectinload(Order.district)
            )
            .where(and_(
                Order.created_at >= start_dt,
                Order.created_at <= end_dt
            ))
            .order_by(desc(Order.created_at))
        )
        result = await session.execute(query)
        orders = result.scalars().all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders Report"
        
        # Headers
        headers = [
            "ID", "Customer", "Phone", "Status", "Total Amount", 
            "Discount", "District", "Delivery Date", "Delivery Time",
            "Items Count", "Created At", "Comment"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for row, order in enumerate(orders, 2):
            ws.cell(row=row, column=1, value=order.id)
            ws.cell(row=row, column=2, value=order.contact_name)
            ws.cell(row=row, column=3, value=order.contact_phone or "")
            ws.cell(row=row, column=4, value=order.status.value if hasattr(order.status, 'value') else str(order.status))
            ws.cell(row=row, column=5, value=order.total_amount)
            ws.cell(row=row, column=6, value=order.discount_amount)
            ws.cell(row=row, column=7, value=order.district.name if order.district else "")
            ws.cell(row=row, column=8, value=order.delivery_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=9, value=order.delivery_time_slot.value if hasattr(order.delivery_time_slot, 'value') else str(order.delivery_time_slot))
            ws.cell(row=row, column=10, value=len(order.items))
            ws.cell(row=row, column=11, value=order.created_at.strftime("%Y-%m-%d %H:%M"))
            ws.cell(row=row, column=12, value=order.comment or "")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"orders_report_{start_date}_{end_date}.xlsx"
        
        return StreamingResponse(
            BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel export not available")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")

@router.get("/orders/{order_id}", response_model=OrderResponse)
async def get_admin_order(
    order_id: int,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get order by ID"""
    query = (
        select(Order)
        .options(
            selectinload(Order.items),
            selectinload(Order.user),
            selectinload(Order.district)
        )
        .where(Order.id == order_id)
    )
    result = await session.execute(query)
    order = result.scalar_one_or_none()
    
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    return order

@router.put("/orders/{order_id}/status", response_model=OrderResponse)
async def update_order_status(
    order_id: int,
    status_update: OrderStatusUpdate,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Update order status"""
    query = select(Order).where(Order.id == order_id)
    result = await session.execute(query)
    order = result.scalar_one_or_none()
    
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    order.status = status_update.status
    order.updated_at = datetime.utcnow()
    
    await session.commit()
    await session.refresh(order, ['items', 'user', 'district'])
    return order

# Districts CRUD
@router.get("/districts", response_model=List[DistrictResponse])
async def get_admin_districts(
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get all districts for admin"""
    query = select(District).order_by(District.name)
    result = await session.execute(query)
    districts = result.scalars().all()
    return districts

@router.get("/districts/{district_id}", response_model=DistrictResponse)
async def get_district_by_id(
    district_id: int,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get district by ID"""
    query = select(District).where(District.id == district_id)
    result = await session.execute(query)
    district = result.scalar_one_or_none()
    
    if not district:
        raise HTTPException(status_code=404, detail="District not found")
    
    return district

@router.post("/districts", response_model=DistrictResponse)
async def create_district(
    district: DistrictCreate,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Create new district"""
    db_district = District(**district.model_dump())
    session.add(db_district)
    await session.commit()
    await session.refresh(db_district)
    return db_district

@router.put("/districts/{district_id}", response_model=DistrictResponse)
async def update_district(
    district_id: int,
    district: DistrictUpdate,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Update district"""
    query = select(District).where(District.id == district_id)
    result = await session.execute(query)
    db_district = result.scalar_one_or_none()
    
    if not db_district:
        raise HTTPException(status_code=404, detail="District not found")
    
    for field, value in district.model_dump(exclude_unset=True).items():
        setattr(db_district, field, value)
    
    await session.commit()
    await session.refresh(db_district)
    return db_district

@router.delete("/districts/{district_id}")
async def delete_district(
    district_id: int,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Delete district"""
    query = select(District).where(District.id == district_id)
    result = await session.execute(query)
    db_district = result.scalar_one_or_none()
    
    if not db_district:
        raise HTTPException(status_code=404, detail="District not found")
    
    await session.delete(db_district)
    await session.commit()
    return {"message": "District deleted successfully"}

# Promo Codes CRUD
@router.get("/promo-codes", response_model=List[PromoCodeResponse])
async def get_admin_promo_codes(
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get all promo codes for admin"""
    query = select(PromoCode).order_by(desc(PromoCode.id))
    result = await session.execute(query)
    promo_codes = result.scalars().all()
    return promo_codes

@router.get("/promo-codes/{promo_code_id}", response_model=PromoCodeResponse)
async def get_promo_code_by_id(
    promo_code_id: int,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get promo code by ID"""
    query = select(PromoCode).where(PromoCode.id == promo_code_id)
    result = await session.execute(query)
    promo_code = result.scalar_one_or_none()
    
    if not promo_code:
        raise HTTPException(status_code=404, detail="Promo code not found")
    
    return promo_code

@router.post("/promo-codes", response_model=PromoCodeResponse)
async def create_promo_code(
    promo_code: PromoCodeCreate,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Create new promo code"""
    try:
        db_promo_code = PromoCode(**promo_code.model_dump())
        session.add(db_promo_code)
        await session.commit()
        await session.refresh(db_promo_code)
        return db_promo_code
    except IntegrityError:
        await session.rollback()
        raise HTTPException(status_code=400, detail="Promo code with this code already exists")

@router.put("/promo-codes/{promo_code_id}", response_model=PromoCodeResponse)
async def update_promo_code(
    promo_code_id: int,
    promo_code: PromoCodeUpdate,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Update promo code"""
    query = select(PromoCode).where(PromoCode.id == promo_code_id)
    result = await session.execute(query)
    db_promo_code = result.scalar_one_or_none()
    
    if not db_promo_code:
        raise HTTPException(status_code=404, detail="Promo code not found")
    
    for field, value in promo_code.model_dump(exclude_unset=True).items():
        setattr(db_promo_code, field, value)
    
    await session.commit()
    await session.refresh(db_promo_code)
    return db_promo_code

@router.delete("/promo-codes/{promo_code_id}")
async def delete_promo_code(
    promo_code_id: int,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Delete promo code"""
    query = select(PromoCode).where(PromoCode.id == promo_code_id)
    result = await session.execute(query)
    db_promo_code = result.scalar_one_or_none()
    
    if not db_promo_code:
        raise HTTPException(status_code=404, detail="Promo code not found")
    
    await session.delete(db_promo_code)
    await session.commit()
    return {"message": "Promo code deleted successfully"}


# Product Packages CRUD
@router.get("/products/{product_id}/packages", response_model=List[ProductPackageResponse])
async def get_product_packages(
    product_id: str,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get all packages for a specific product"""
    query = (
        select(ProductPackage)
        .where(ProductPackage.product_id == product_id)
        .order_by(ProductPackage.sort_order, ProductPackage.id)
    )
    result = await session.execute(query)
    packages = result.scalars().all()
    return packages


@router.post("/products/{product_id}/packages", response_model=ProductPackageResponse, status_code=201)
async def create_product_package(
    product_id: str,
    package: ProductPackageCreate,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Create a new package for a product"""
    # Check if product exists
    product_query = select(Product).where(Product.id == product_id)
    product_result = await session.execute(product_query)
    if not product_result.scalar_one_or_none():
        raise HTTPException(status_code=404, detail="Product not found")
    
    # Check if package_id already exists for this product
    existing_query = select(ProductPackage).where(
        ProductPackage.product_id == product_id,
        ProductPackage.package_id == package.package_id
    )
    existing_result = await session.execute(existing_query)
    if existing_result.scalar_one_or_none():
        raise HTTPException(
            status_code=400, 
            detail=f"Package with ID '{package.package_id}' already exists for this product"
        )
    
    try:
        db_package = ProductPackage(
            product_id=product_id,
            **package.model_dump()
        )
        session.add(db_package)
        await session.commit()
        await session.refresh(db_package)
        return db_package
    except IntegrityError:
        await session.rollback()
        raise HTTPException(status_code=400, detail="Failed to create package")


@router.get("/products/{product_id}/packages/{package_id}", response_model=ProductPackageResponse)
async def get_product_package(
    product_id: str,
    package_id: int,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Get a specific package by ID"""
    query = select(ProductPackage).where(
        ProductPackage.product_id == product_id,
        ProductPackage.id == package_id
    )
    result = await session.execute(query)
    package = result.scalar_one_or_none()
    
    if not package:
        raise HTTPException(status_code=404, detail="Package not found")
    
    return package


@router.put("/products/{product_id}/packages/{package_id}", response_model=ProductPackageResponse)
async def update_product_package(
    product_id: str,
    package_id: int,
    package: ProductPackageUpdate,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Update a product package"""
    query = select(ProductPackage).where(
        ProductPackage.product_id == product_id,
        ProductPackage.id == package_id
    )
    result = await session.execute(query)
    db_package = result.scalar_one_or_none()
    
    if not db_package:
        raise HTTPException(status_code=404, detail="Package not found")
    
    for field, value in package.model_dump(exclude_unset=True).items():
        setattr(db_package, field, value)
    
    await session.commit()
    await session.refresh(db_package)
    return db_package


@router.delete("/products/{product_id}/packages/{package_id}", status_code=204)
async def delete_product_package(
    product_id: str,
    package_id: int,
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Delete a product package"""
    query = select(ProductPackage).where(
        ProductPackage.product_id == product_id,
        ProductPackage.id == package_id
    )
    result = await session.execute(query)
    db_package = result.scalar_one_or_none()
    
    if not db_package:
        raise HTTPException(status_code=404, detail="Package not found")
    
    # Delete image from S3 if exists
    if db_package.image_url:
        try:
            await s3_service.delete_image(db_package.image_url)
        except Exception as e:
            print(f"Warning: Failed to delete package image: {e}")
    
    await session.delete(db_package)
    await session.commit()


@router.post("/products/{product_id}/packages/{package_id}/image", response_model=ImageUploadResponse)
async def upload_package_image(
    product_id: str,
    package_id: int,
    image: UploadFile = File(...),
    session: AsyncSession = Depends(get_async_session),
    current_admin: AdminUser = Depends(get_current_admin)
):
    """Upload image for a product package"""
    # Check if package exists
    query = select(ProductPackage).where(
        ProductPackage.product_id == product_id,
        ProductPackage.id == package_id
    )
    result = await session.execute(query)
    db_package = result.scalar_one_or_none()
    
    if not db_package:
        raise HTTPException(status_code=404, detail="Package not found")
    
    try:
        # Delete old image if exists
        if db_package.image_url:
            await s3_service.delete_image(db_package.image_url)
        
        # Upload new image
        image_url = await s3_service.upload_image(image, folder="packages")
        
        # Update package with new image URL
        db_package.image_url = image_url
        await session.commit()
        
        return ImageUploadResponse(url=image_url)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))